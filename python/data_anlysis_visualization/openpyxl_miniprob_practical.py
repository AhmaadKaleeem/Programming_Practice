#Create a workbook with 3 sheets (Students, Teachers, Subjects).
# In Students, add 3 rows of names and ages.
# In Teachers, format the headers in bold blue font.
# In Subjects, write a list of 5 subjects.

# loading librares for excel automation
import os
from openpyxl import Workbook , load_workbook 
from openpyxl.styles import  Font 

wb = Workbook()
# Creating three worksheets

sheet1 = wb.active
sheet2 = wb.create_sheet(title ="Teachers")
sheet3 = wb.create_sheet(title ="Subjects")

sheet1.title = "Students"


studentdata = [('Ali',19),('Ahmad',20),('Hanif',19)]
sheet1.append(['Name','Age'])
for row in studentdata:
  sheet1.append(row)

sheet2['A1'] = "Teachers"
sheet2['A1'].font = Font(bold = True, color = "0000FF")

subjects = ['OOP','DB','AI','M:','DS']
sheet3.append(row)
# 2nd approch
# sheet1.append(['Ali',19])
# sheet1.append(['Ahmad',20])
# sheet1.append(['Hanif',19])

wb.save("Project1.xlsx")
print('Filed Saved Successfully!')
