from openpyxl import Workbook
# Writing sample data for excel automation
wb= Workbook()
ws = wb.active
ws.title = "Data"
ws['A1'] = "Name"
ws['B1'] = "Age"

ws.append(['Ahmad',20])
ws.append(['Kaleem',25])

wb.save("Data.xlsx")